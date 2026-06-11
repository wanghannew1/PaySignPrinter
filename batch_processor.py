"""
Batch processing module for DingTalk approval attachments.
Handles: download, signature insertion, and printing.
"""

import json
import platform
import subprocess
import tempfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

import cache_manager
from logger_config import logger


def _load_role_mapping():
    mapping_path = Path(__file__).parent / "role_mapping.json"
    if mapping_path.exists():
        try:
            with open(mapping_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError) as e:
            print(f"加载角色映射失败: {e}，使用内置默认值")
    return {
        "五险一金": "业务审核",
        "业务审核": "业务审核",
        "部门负责人": "部长签字",
        "部长签字": "部长签字",
        "分管副总签字": "部长签字",
        "财务": "财务审核",
        "财务审核": "财务审核",
        "总经理": "总经理签字",
        "总经理签字": "总经理签字",
    }


def get_approver_role(show_name: str) -> Optional[str]:
    show_name_lower = show_name.lower()
    mapping = _load_role_mapping()
    for role_keyword, sig_keyword in mapping.items():
        if role_keyword in show_name_lower:
            return sig_keyword
    return None


def is_approval_passed(details: dict) -> Tuple[bool, str]:
    """Check if approval is passed (all required approvers agreed)."""
    records = details.get("operationRecords", [])
    for record in records:
        if record.get("type") in ("EXECUTE_TASK_NORMAL", "EXECUTE_TASK_TRANSFER"):
            result = record.get("result", "NONE")
            if result == "REFUSE":
                return False, f"审批被拒绝: {record.get('showName', '')}"
            elif result == "BACK":
                return False, f"审批被退回: {record.get('showName', '')}"
    return True, "审批通过"


def get_approvers_with_roles(details: dict) -> List[Dict]:
    records = details.get("operationRecords", [])
    approvers = []
    for record in records:
        if record.get("type") in ("EXECUTE_TASK_NORMAL", "EXECUTE_TASK_TRANSFER"):
            if record.get("result") == "AGREE":
                show_name = record.get("showName", "")
                role = get_approver_role(show_name)
                if role:
                    approvers.append({
                        "userId": record.get("userId"),
                        "showName": show_name,
                        "role": role,
                    })
    return approvers


def find_signature_image(user_id: str, signatures_dir: Path) -> Optional[Path]:
    """Find signature image for a user."""
    if not user_id:
        return None
    sig_path = signatures_dir / f"{user_id}.png"
    if sig_path.exists():
        return sig_path
    for f in signatures_dir.glob("*.png"):
        if user_id in f.name:
            return f
    return None


def _is_cell_in_merged_range(ws, row, col):
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
            return merged_range
    return None


def _estimate_text_span(text, col_width):
    if not col_width or col_width == 0:
        col_width = 8.43
    width = sum(2 if '\u4e00' <= c <= '\u9fff' else 1 for c in str(text))
    return width / col_width


def _split_merged_for_text(ws, row, col):
    from openpyxl.utils import get_column_letter

    merged = _is_cell_in_merged_range(ws, row, col)
    if not merged:
        cell = ws.cell(row=row, column=col)
        if cell.value and "部长、分管副总签字" in str(cell.value):
            cell.value = str(cell.value).replace("部长、分管副总签字", "部长签字")
        return col + 1

    cell = ws.cell(row=row, column=col)
    text = str(cell.value) if cell.value else ""

    if "部长、分管副总签字" in text:
        cell.value = text.replace("部长、分管副总签字", "部长签字")
        needed_cols = 3
    else:
        needed_cols = 2

    total_cols = merged.max_col - merged.min_col + 1
    if needed_cols >= total_cols:
        return merged.max_col + 1

    merged_str = str(merged)
    ws.unmerge_cells(merged_str)

    new_end = merged.min_col + needed_cols - 1
    ws.merge_cells(start_row=row, start_column=merged.min_col,
                   end_row=row, end_column=new_end)

    for c in range(new_end + 1, merged.max_col + 1):
        ws.cell(row=row, column=c).value = None

    return new_end + 1


def find_all_signature_positions(ws) -> Dict[str, Tuple[int, int]]:
    """Find all signature positions in the worksheet."""
    positions = {}
    keyword_groups = [
        (["总经理签字"], "总经理签字"),
        (["部长签字", "部长、分管副总签字", "分管副总签字"], "部长签字"),
        (["财务审核"], "财务审核"),
        (["业务审核"], "业务审核"),
    ]

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                value = str(cell.value).strip()
                for texts, role_key in keyword_groups:
                    if role_key not in positions:
                        for text in texts:
                            if text in value:
                                positions[role_key] = (row, col)
                                break
    return positions


def _extract_first_row_title(ws) -> Optional[str]:
    """Find the first non-empty cell in row 1 as the table title."""
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value:
            return str(cell.value).strip()
    return None


def _build_output_path(excel_path: Path, output_path: Path, ws) -> Path:
    """If the original file is an unrenamed export (tddd_dialog*), append row-1 title."""
    original_name = excel_path.name
    if original_name.lower().startswith("tddd_dialog"):
        title = _extract_first_row_title(ws)
        if title:
            safe_title = sanitize_dir_name(title)
            new_name = f"signed_{Path(original_name).stem}-{safe_title}.xlsx"
            return output_path.parent / new_name
    return output_path


def _find_total_row(ws) -> int:
    keywords = ["合计", "总计", "合计金额", "合计费用", "合计支付"]
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                val = str(cell.value).strip()
                for kw in keywords:
                    if kw in val:
                        return row
    return 0


def _apply_border_styles(ws, signature_positions):
    from openpyxl.styles import Border, Side
    try:
        thick = Side(style='medium', color='000000')
        none = Side(style=None)

        total_row = _find_total_row(ws)
        sig_rows = sorted({r for (r, _) in signature_positions.values()})
        if not sig_rows:
            logger.info("[BORDER] 无签名位置，跳过边框设置")
            return
        sig_start = min(sig_rows)
        sig_end = max(sig_rows)
        last_col = ws.max_column

        border_start = 3
        border_end = sig_end + 2

        if border_start <= border_end:
            for r in range(border_start, border_end + 1):
                for c in range(1, last_col + 1):
                    is_top = r == border_start
                    is_bottom = r == border_end
                    is_left = c == 1
                    is_right = c == last_col
                    if not (is_top or is_bottom or is_left or is_right):
                        continue
                    cell = ws.cell(row=r, column=c)
                    cell.border = Border(
                        top=thick if is_top else cell.border.top,
                        bottom=thick if is_bottom else cell.border.bottom,
                        left=thick if is_left else cell.border.left,
                        right=thick if is_right else cell.border.right,
                    )

        logger.info(f"[BORDER] 外框从第{border_start}行到第{border_end}行，内部单元格边框保留")
    except Exception as e:
        logger.warning(f"[BORDER] 边框设置出错: {e}")


def _hide_columns(ws):
    headers_to_hide = {"部门", "岗位", "职工号"}
    header_row = 3
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        col_letter = cell.column_letter
        hidden = ws.column_dimensions[col_letter].hidden
        if cell.value and str(cell.value).strip() in headers_to_hide:
            hidden = True
        ws.column_dimensions[col_letter].hidden = hidden


def adjust_excel_for_print(ws, signature_positions=None) -> None:
    """
    调整 Excel 打印设置：横向打印，A4 纸，左边距 2cm，其他边距 1cm，
    所有列缩放到 1 页宽，水平居中。
    在嵌入签名图片前调用。
    """
    try:
        ws.page_setup.paperSize = 9          # A4
        ws.page_setup.orientation = "landscape"
        ws.sheet_properties.pageSetUpPr.fitToPage = True  # 正确路径，必须先启用
        ws.page_setup.fitToWidth = 1         # 缩放到 1 页宽
        ws.page_setup.fitToHeight = 0        # 高度自动分页
        ws.page_margins.left = 0.8           # 2cm
        ws.page_margins.right = 0.4          # 1cm
        ws.page_margins.top = 0.4            # 1cm
        ws.page_margins.bottom = 0.4         # 1cm
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = False
        ws.print_options.gridLines = False
        logger.info("[PRINT] 已调整: 横向A4, 左2cm其余1cm, 1页宽, fitToPage=True, 网格线关闭")

        _hide_columns(ws)

        if signature_positions:
            _apply_border_styles(ws, signature_positions)
    except Exception as e:
        logger.warning(f"[PRINT] 调整打印设置时出错: {e}")


def _insert_signature_to_excel_openpyxl(
    excel_path: Path,
    approvers: List[Dict],
    signatures_dir: Path,
    output_path: Path,
) -> Tuple[bool, List[str], Path]:
    inserted_roles = []
    try:
        logger.info(f"[SIGN] Loading workbook: {excel_path.name}")
        wb = load_workbook(str(excel_path))
        ws = wb.active

        positions = find_all_signature_positions(ws)
        adjust_excel_for_print(ws, positions)
        logger.info(f"[SIGN] Found positions: {positions}")
        if not positions:
            logger.warning(f"[SIGN] No signature positions found in {excel_path.name}")
            return False, [], output_path

        logger.info(f"[SIGN] Approvers: {[a['role'] for a in approvers]}")
        for approver in approvers:
            role = approver.get("role")
            user_id = approver.get("userId")
            if not role or not user_id:
                logger.info(f"[SIGN] Skipping approver: missing role or userId")
                continue

            if role not in positions:
                logger.info(f"[SIGN] Role '{role}' not in positions")
                continue

            sig_path = find_signature_image(user_id, signatures_dir)
            if not sig_path:
                logger.warning(f"[SIGN] Signature image not found for user {user_id}")
                continue

            row, col = positions[role]
            target_col = _split_merged_for_text(ws, row, col)

            img = XLImage(str(sig_path))
            img.width = 120
            img.height = 60
            cell_addr = f"{ws.cell(row=row, column=target_col).coordinate}"
            ws.add_image(img, cell_addr)
            inserted_roles.append(role)
            logger.info(f"[SIGN] Inserted signature for {role} at {cell_addr}")

        actual_output = _build_output_path(excel_path, output_path, ws)
        wb.save(str(actual_output))
        logger.info(f"[SIGN] Saved to {actual_output.name}, inserted: {inserted_roles}")
        return len(inserted_roles) > 0, inserted_roles, actual_output
    except Exception as e:
        logger.error(f"[SIGN] Insertion failed: {e}")
        return False, [], output_path


def _convert_xls_to_xlsx_libreoffice(xls_path: Path) -> Optional[Path]:
    """Convert .xls to .xlsx using LibreOffice (cross-platform, no Excel required)."""
    xlsx_path = xls_path.with_suffix(".xlsx")
    try:
        import subprocess
        result = subprocess.run(
            [
                "soffice",
                "--headless",
                "--convert-to", "xlsx",
                str(xls_path.resolve()),
                "--outdir", str(xls_path.parent.resolve()),
            ],
            capture_output=True,
            text=True,
            timeout=60,
        )
        if result.returncode == 0 and xlsx_path.exists():
            return xlsx_path
        else:
            print(f"LibreOffice转换失败: {result.stderr}")
            return None
    except FileNotFoundError:
        print("LibreOffice未安装或未在PATH中")
        return None
    except Exception as e:
        print(f"LibreOffice转换异常: {e}")
        return None


def _convert_xls_to_xlsx_windows(xls_path: Path) -> Optional[Path]:
    """Try Excel COM first, fallback to LibreOffice."""
    try:
        import pythoncom
        import win32com.client
        from win32com.client import constants

        pythoncom.CoInitialize()
        excel = None
        wb = None

        try:
            xlsx_path = xls_path.with_suffix(".xlsx")
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            wb = excel.Workbooks.Open(str(xls_path.resolve()))
            # Use numeric format code 51 for xlsx (Excel 2007+ format)
            # WPS may not recognize constants.xlOpenXMLWorkbook
            wb.SaveAs(str(xlsx_path.resolve()), FileFormat=51)
            wb.Close(SaveChanges=False)
            excel.Quit()

            return xlsx_path
        except Exception as e:
            print(f"WPS/Excel COM转换失败: {e}，尝试不带FileFormat...")
            try:
                # Fallback: save without explicit format, let WPS infer from extension
                xlsx_path = xls_path.with_suffix(".xlsx")
                wb = excel.Workbooks.Open(str(xls_path.resolve()))
                wb.SaveAs(str(xlsx_path.resolve()))
                wb.Close(SaveChanges=False)
                excel.Quit()
                return xlsx_path
            except Exception as e2:
                print(f"不带FileFormat也失败: {e2}，尝试LibreOffice...")
                return _convert_xls_to_xlsx_libreoffice(xls_path)
        finally:
            try:
                if wb:
                    wb.Close(SaveChanges=False)
                if excel:
                    excel.Quit()
            except Exception:
                pass
            pythoncom.CoUninitialize()
    except ImportError:
        print("pywin32未安装，尝试LibreOffice...")
        return _convert_xls_to_xlsx_libreoffice(xls_path)


def _insert_signature_to_excel_windows(
    excel_path: Path,
    approvers: List[Dict],
    signatures_dir: Path,
    output_path: Path,
) -> Tuple[bool, List[str], Path]:
    xlsx_path = _convert_xls_to_xlsx_windows(excel_path)
    if xlsx_path is None:
        return False, [], output_path

    try:
        success, inserted, actual_output = _insert_signature_to_excel_openpyxl(
            xlsx_path, approvers, signatures_dir, output_path
        )
        return success, inserted, actual_output
    finally:
        try:
            xlsx_path.unlink()
        except Exception:
            pass


def insert_signature_to_excel(
    excel_path: Path,
    approvers: List[Dict],
    signatures_dir: Path,
    output_path: Path,
) -> Tuple[bool, List[str], Path]:
    ext = excel_path.suffix.lower()
    if platform.system() == "Windows" and ext == ".xls":
        return _insert_signature_to_excel_windows(
            excel_path, approvers, signatures_dir, output_path
        )
    return _insert_signature_to_excel_openpyxl(
        excel_path, approvers, signatures_dir, output_path
    )


def _print_with_com(file_path: Path, printer_name: Optional[str] = None) -> bool:
    """Print using WPS/Excel COM (Windows only)."""
    if platform.system() != "Windows":
        return False
    try:
        import pythoncom
        import win32com.client

        pythoncom.CoInitialize()
        app = None
        wb = None

        try:
            app = win32com.client.Dispatch("Excel.Application")
            app.Visible = False
            app.DisplayAlerts = False

            wb = app.Workbooks.Open(str(file_path.resolve()))
            if printer_name:
                wb.PrintOut(ActivePrinter=printer_name)
            else:
                wb.PrintOut()
            wb.Close(SaveChanges=False)
            app.Quit()
            return True
        except Exception as e:
            print(f"WPS/Excel打印失败: {e}")
            return False
        finally:
            try:
                if wb:
                    wb.Close(SaveChanges=False)
                if app:
                    app.Quit()
            except Exception:
                pass
            pythoncom.CoUninitialize()
    except ImportError:
        return False


def _print_with_libreoffice(file_path: Path, printer_name: Optional[str] = None) -> bool:
    """Print using LibreOffice (cross-platform fallback)."""
    try:
        if file_path.suffix.lower() in (".xlsx", ".xls", ".docx", ".doc"):
            cmd = [
                "soffice",
                "--headless",
                "--print",
                str(file_path),
            ]
            if printer_name:
                cmd.extend(["--printer", printer_name])
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=60,
            )
            return result.returncode == 0
        return False
    except FileNotFoundError:
        print("LibreOffice未安装")
        return False


def print_file(file_path: Path, printer_name: Optional[str] = None) -> bool:
    """Print file. Windows uses WPS/Excel COM, Linux uses LibreOffice."""
    if platform.system() == "Windows":
        success = _print_with_com(file_path, printer_name)
        if success:
            return True
        print("COM打印失败，尝试LibreOffice...")
        return _print_with_libreoffice(file_path, printer_name)
    else:
        return _print_with_libreoffice(file_path, printer_name)


def sanitize_dir_name(name: str) -> str:
    """Sanitize string for use as directory name."""
    invalid_chars = '\\/:*?"<>|'
    for char in invalid_chars:
        name = name.replace(char, "_")
    return name.strip()


def process_single_approval(
    instance_id: str,
    token: str,
    signatures_dir: Path,
    output_base_dir: Path,
    dingtalk_api_module,
) -> dict:
    """Process a single approval: download, sign, print."""
    result = {
        "instance_id": instance_id,
        "success": False,
        "message": "",
        "downloaded": [],
        "signed": [],
        "signed_files": [],
        "skipped": False,
        "title": "",
    }

    try:
        logger.info(f"[BATCH] Getting details for {instance_id[:20]}...")
        details = cache_manager.get_cached_instance_details(instance_id)
        if details is None:
            logger.info(f"[BATCH] Cache miss, calling get_instance_details API...")
            details = dingtalk_api_module.get_instance_details(instance_id, token)
            cache_manager.cache_instance_details(instance_id, details)
            logger.info(f"[BATCH] Got details from API")
        else:
            logger.info(f"[BATCH] Cache hit for details")
    except Exception as e:
        logger.error(f"[BATCH] Failed to get details: {e}")
        result["message"] = f"获取详情失败: {e}"
        return result

    business_id = details.get("businessId", instance_id[:20])
    result["business_id"] = business_id

    passed, msg = is_approval_passed(details)
    if not passed:
        result["skipped"] = True
        result["message"] = msg
        return result

    approvers = get_approvers_with_roles(details)
    if not approvers:
        result["message"] = "未找到可签名的审批角色"
        return result

    form_values = details.get("formComponentValues", [])
    attachments = dingtalk_api_module.extract_attachments(form_values)

    if not attachments:
        result["message"] = "无附件"
        return result

    instance_dir = output_base_dir / sanitize_dir_name(business_id)
    instance_dir.mkdir(parents=True, exist_ok=True)

    for att in attachments:
        file_id = att.get("fileId")
        file_name = att.get("fileName", "unknown")
        file_type = att.get("fileType", "")

        if "汇总表" in file_name:
            logger.info(f"[BATCH] Skipping summary table: {file_name}")
            continue

        try:
            logger.info(f"[BATCH] Processing attachment: {file_name}")
            download_url = cache_manager.get_cached_download_url(instance_id, file_id)
            if download_url is None:
                logger.info(f"[BATCH] Download URL cache miss, calling API...")
                download_url, needs_rename = dingtalk_api_module.get_download_url(
                    instance_id, file_id, token
                )
                cache_manager.cache_download_url(instance_id, file_id, download_url)
                logger.info(f"[BATCH] Got download URL")
            else:
                logger.info(f"[BATCH] Download URL cache hit")
            
            logger.info(f"[BATCH] Downloading file bytes...")
            file_bytes = dingtalk_api_module.download_file_bytes(download_url)
            logger.info(f"[BATCH] Downloaded {len(file_bytes)} bytes")

            # Save to instance dir directly (no nested subdir)
            safe_name = sanitize_dir_name(file_name)
            file_path = instance_dir / safe_name
            counter = 1
            while file_path.exists():
                stem = Path(safe_name).stem
                suffix = Path(safe_name).suffix
                file_path = instance_dir / f"{stem}_{counter}{suffix}"
                counter += 1

            with open(file_path, "wb") as f:
                f.write(file_bytes)

            result["downloaded"].append(file_name)

            # If Excel, insert signatures
            if file_type in ("xlsx", "xls") or file_path.suffix in (".xlsx", ".xls"):
                logger.info(f"[BATCH] Inserting signatures into {file_name}...")
                signed_name = f"signed_{file_path.stem}.xlsx"
                signed_path = instance_dir / signed_name
                success, inserted, actual_signed_path = insert_signature_to_excel(
                    file_path, approvers, signatures_dir, signed_path
                )
                if success:
                    logger.info(f"[BATCH] Signature insertion success: {inserted}")
                    result["signed"].extend(inserted)
                    result["signed_files"].append(str(actual_signed_path))
                else:
                    logger.warning(f"[BATCH] Signature insertion failed for {file_name}")
            else:
                logger.info(f"[BATCH] Non-Excel file, skipping signature: {file_name}")

        except Exception as e:
            logger.error(f"[BATCH] Error processing {file_name}: {e}")
            continue

    result["success"] = True
    result["message"] = (
        f"下载 {len(result['downloaded'])} 个, "
        f"签名 {len(result['signed'])} 处"
    )
    return result
